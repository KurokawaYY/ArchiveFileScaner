import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import logging


class ExcelExporter:
    def __init__(self):
        self.logger = logging.getLogger(__name__)

        # ⬅️ ЦВЕТА ДЛЯ EXCEL
        self.colors = {
            'file': 'FFFFFF',  # Белый
            'directory': 'E3F2FD',  # Голубой
            'archive': 'FFF3E0',  # Оранжевый
            'archive_file': 'F3E5F5',  # Фиолетовый
            'archive_directory': 'E8F5E8',  # Зеленый
            'nested_archive': 'FFEBEE',  # Красный
            'archive_error': 'FFCDD2'  # Розовый
        }

    def export_to_excel(self, data: List[Dict[str, Any]], file_path: str):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Структура файлов"

            # ⬅️ УБРАЛИ КОЛОНКУ "№"
            headers = [
                "Тип", "Имя", "Путь",
                "Размер\n(байты)", "Глубина", "Уровень\nархива",
                "Контрольная\nсумма"
            ]
            ws.append(headers)

            self._apply_header_styles(ws, headers)

            # ⬅️ СОРТИРУЕМ ДАННЫЕ ДЛЯ ПРАВИЛЬНОЙ ИЕРАРХИИ
            sorted_data = sorted(data, key=lambda x: x['path'])

            for item in sorted_data:
                # ⬅️ ДОБАВЛЯЕМ ОТСТУПЫ ДЛЯ ИМЕНИ
                indented_name = self._format_name_with_indent(item)

                row = [
                    self._get_type_display(item['type']),  # ⬅️ РУССКОЕ НАЗВАНИЕ
                    indented_name,
                    item['path'],
                    item.get('size', 0),
                    item['depth'],
                    item.get('archive_depth', 0),
                    item.get('crc32', 'Н/Д')
                ]
                ws.append(row)

                # ⬅️ ПРИМЕНЯЕМ ЦВЕТ ДЛЯ СТРОКИ
                self._apply_row_style(ws, ws.max_row, item['type'])

            self._apply_data_styles(ws, len(data))
            self._auto_adjust_columns(ws)
            ws.freeze_panes = 'A2'
            self._add_explanations(ws, len(data) + 4)

            wb.save(file_path)
            self.logger.info(f"Файл успешно сохранен: {file_path}")

        except Exception as e:
            self.logger.error(f"Ошибка при экспорте в Excel: {e}")
            raise

    def _format_name_with_indent(self, item):
        """Форматирует имя с отступами для иерархии"""
        depth = item['depth'] + item.get('archive_depth', 0)

        # ⬅️ ДОПОЛНИТЕЛЬНЫЕ ОТСТУПЫ ДЛЯ ФАЙЛОВ В ПАПКАХ
        additional_indent = 0
        if item['type'] in ['archive_file', 'file']:
            # Проверяем, находится ли файл внутри папки
            if '/' in item['path'] or '\\' in item['path']:
                additional_indent = 1

        total_indent = depth + additional_indent
        indent = "    " * total_indent

        # ⬅️ ПРЕФИКСЫ ДЛЯ РАЗНЫХ ТИПОВ
        prefix = self._get_name_prefix(item['type'])

        return f"{indent}{prefix}{item['name']}"

    def _get_name_prefix(self, item_type):
        """Возвращает префикс для имени"""
        prefixes = {
            'file': '[ФАЙЛ] ',
            'directory': '[ПАПКА] ',
            'archive': '[АРХИВ] ',
            'archive_file': '[ФАЙЛ В АРХИВЕ] ',
            'archive_directory': '[ПАПКА В АРХИВЕ] ',
            'nested_archive': '[ВЛОЖ. АРХИВ] ',
            'archive_error': '[ОШИБКА АРХИВА] '
        }
        return prefixes.get(item_type, '')

    def _get_type_display(self, item_type: str) -> str:
        """Возвращает русское название типа"""
        type_map = {
            'file': 'Файл',
            'directory': 'Папка',
            'archive': 'Архив',
            'archive_file': 'Файл в архиве',
            'archive_directory': 'Папка в архиве',  # ⬅️ РУССКОЕ НАЗВАНИЕ
            'nested_archive': 'Влож. архив',
            'archive_error': 'Ошибка архива'
        }
        return type_map.get(item_type, item_type)

    def _apply_row_style(self, worksheet, row_num, item_type):
        """Применяет цвет к строке"""
        color = self.colors.get(item_type, 'FFFFFF')
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # ⬅️ ТЕПЕРЬ 7 СТОЛБЦОВ ВМЕСТО 8
        for col in range(1, 8):
            cell = worksheet.cell(row=row_num, column=col)
            cell.fill = fill

    def _apply_header_styles(self, worksheet, headers):
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(bottom=Side(style='thin'))

        for col in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _apply_data_styles(self, worksheet, data_count: int):
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(2, data_count + 2):
            # ⬅️ ТЕПЕРЬ 7 СТОЛБЦОВ
            for col in range(1, 8):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border

                # Для имен - выравнивание по левому краю без переноса
                if col == 2:  # Столбец "Имя" (теперь второй)
                    cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=False)
                else:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

    def _auto_adjust_columns(self, worksheet):
        # ⬅️ ОБНОВЛЕННЫЕ ШИРИНЫ КОЛОНОК (7 ВМЕСТО 8)
        column_widths = {
            1: 15,  # Тип
            2: 65,  # Имя (с отступами)
            3: 85,  # Путь
            4: 13,  # Размер
            5: 10,  # Глубина
            6: 10,  # Уровень архива
            7: 25  # Контрольная сумма
        }

        for col_num, width in column_widths.items():
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = width

        worksheet.row_dimensions[1].height = 40

    def _add_explanations(self, worksheet, start_row: int):
        explanations = [
            "Пояснения:",
            "• Контрольная сумма - техническая характеристика файла для проверки целостности",
            "• CRC32: XXXXXXXX - уникальный идентификатор содержимого файла",
            "• Пустой файл - файл размером 0 байт",
            "• [ВЛ.АРХ] - вложенный архив (требуется 7-Zip для полного сканирования)",
            "• Требуется 7-Zip - установите 7-Zip для работы с некоторыми архивами",
            "• Отступы показывают иерархию вложенности файлов и папок",
            "",
            "Цветовое кодирование:",
            "• Белый - обычные файлы",
            "• Голубой - папки",
            "• Оранжевый - архивы",
            "• Фиолетовый - файлы в архивах",
            "• Зеленый - папки в архивах",
            "• Красный - вложенные архивы",
            "• Розовый - ошибки архивов"
        ]

        for i, explanation in enumerate(explanations):
            cell = worksheet.cell(row=start_row + i, column=2)  # ⬅️ Теперь столбец B для пояснений
            cell.value = explanation
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            if i == 0:
                cell.font = Font(bold=True, color="366092", size=12)
            elif i == 1:
                cell.font = Font(italic=True, color="666666")
            elif i >= 8:  # Цветовое кодирование
                cell.font = Font(color="FF6600", italic=True)